import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from io import BytesIO

# Streamlit app title
st.title("Employee Summary Sheet Generator")
st.write(
    "Upload a master Excel file containing all employee sheets. The app will generate a summary sheet with formulas linking back to each source sheet and add it to the original file for download."
)

# Upload file
uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file is not None:
    try:
        # Load workbook
        wb = load_workbook(filename=uploaded_file)

        # Exclude hidden sheets and any summary-like sheets
        employee_sheets = [
            s
            for s in wb.sheetnames
            if "summary" not in s.lower() and wb[s].sheet_state == "visible"
        ]

        # Define columns
        columns = [
            "S.No",
            "Name",
            "Father Name",
            "CNIC #",
            "Designation",
            "Joining Date",
            "Salary",
            "Overtime",
            "Total",
            "Sheet Name",
        ]

        # Add summary sheet
        if "Summary" in wb.sheetnames:
            del wb["Summary"]
        ws_summary = wb.create_sheet("Summary")

        # Write headers
        for col_num, col_name in enumerate(columns, start=1):
            ws_summary.cell(row=1, column=col_num, value=col_name)

        # Fill summary with formulas referencing each employee sheet
        for idx, sheet_name in enumerate(employee_sheets, start=2):
            ws_summary.cell(row=idx, column=1, value=idx - 1)  # S.No

            # Add formulas for Name, Father Name, CNIC, Designation, etc.
            ws_summary.cell(row=idx, column=2, value=f"='{sheet_name}'!G6")
            ws_summary.cell(row=idx, column=3, value=f"='{sheet_name}'!G7")
            ws_summary.cell(row=idx, column=4, value=f"='{sheet_name}'!G8")
            ws_summary.cell(row=idx, column=5, value=f"='{sheet_name}'!G9")
            ws_summary.cell(row=idx, column=6, value=f"='{sheet_name}'!N3")
            ws_summary.cell(row=idx, column=7, value=f"='{sheet_name}'!L38")
            ws_summary.cell(row=idx, column=8, value=f"='{sheet_name}'!L39")
            ws_summary.cell(row=idx, column=9, value=f"='{sheet_name}'!L40")
            ws_summary.cell(row=idx, column=10, value=sheet_name)

        # Save updated workbook to memory
        output = BytesIO()
        wb.save(output)

        st.success("✅ Summary sheet with links added to the uploaded workbook!")
        st.download_button(
            label="Download Updated Excel File",
            data=output.getvalue(),
            file_name="Updated_Employee_File.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"❌ Error processing file: {e}")
